"""
Обработка запросов
"""
import base64
import json
import os
import PyPDF2
import openpyxl
import fpdf
import db


def __ids_to_names(user, data):
    for region in data['regions']:
        if user['region_id'] == region['id']:
            user['region_id'] = region['region_name']
    for city in data['cities']:
        if user['city_id'] == city['id']:
            user['city_id'] = city['city_name']
    return user


def __name_to_id(user, data):
    for region in data['regions']:
        if user['region_id'] == region['region_name']:
            user['region_id'] = region['id']
    for city in data['cities']:
        if user['city_id'] == city['city_name']:
            user['city_id'] = city['id']
    return user


def __replace_invalid(user, data):
    if user['region_id'] not in [item['id'] for item in data['regions']]:
        user['region_id'] = ''
    if user['city_id'] not in [item['id'] for item in data['cities']]:
        user['city_id'] = ''
    for k in user:
        if user[k] is None:
            user[k] = ''
    return user


def __apply(functions: list, *args):
    for func in functions:
        func(*args)


def __update(conn, columns, values, data):
    try:
        db.insert(conn, 'users', columns, list(values.values()))
    except:
        db.insert(conn, 'users', columns, data['users'])
        return json.dumps({'body': 'ERR'})


def users():
    conn = db.connect('../db.sqlite')
    data = {table: db.select(conn, table) for table in ['users', 'regions', 'cities']}

    for i in range(len(data['users'])):
        data['users'][i] = __replace_invalid(data['users'][i], data)

    result = json.dumps({'body': data['users']})
    conn.close()
    return result


def regions():
    conn = db.connect('../db.sqlite')
    result = json.dumps({'body': db.select(conn, 'regions')})
    conn.close()
    return result


def cities(endpoint: str = None):
    cond = ''
    try:
        cond = endpoint.split('/')[1].split('=')[1]
    except IndexError:
        pass
    conn = db.connect('../db.sqlite')
    result = json.dumps({'body': db.select(conn, 'cities', f'cities.region_id = {cond}')})
    conn.close()
    return result


def add_user(data):
    conn = db.connect('../db.sqlite')
    db.insert(conn, 'users', data.keys(), data.values())
    conn.close()
    return json.dumps({'body': 'OK'})


def import_excel(data):
    tmp_path = 'tmp.xlsx'

    with open(tmp_path, 'wb') as f:
        f.write(data)

    try:
        workbook = openpyxl.load_workbook(tmp_path)
    except:
        os.remove(tmp_path)
        return json.dumps({'body': 'ERR'})
    os.remove(tmp_path)

    conn = db.connect('../db.sqlite')
    db.delete_from(conn, 'users')

    columns = db.get_columns(conn, 'users')[1:]
    data = {table: db.select(conn, table) for table in ['users', 'regions', 'cities']}

    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, min_col=2):
        values = {}
        for i, cell in enumerate(row):
            values[columns[i]] = cell.value
        __apply([__name_to_id, __replace_invalid], values, data)
        __update(conn, columns, values, data)

    conn.close()
    return json.dumps({'body': 'OK'})


def export_excel():
    tmp_path = 'tmp.xlsx'

    conn = db.connect('../db.sqlite')
    columns = db.get_columns(conn, 'users')
    data = {table: db.select(conn, table) for table in ['users', 'regions', 'cities']}
    conn.close()

    for i in range(len(data['users'])):
        __ids_to_names(data['users'][i], data)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(columns)
    for row in data['users']:
        sheet.append(list(row.values()))

    workbook.save(tmp_path)
    with open(tmp_path, 'rb') as f:
        data = f.read()
        f.close()
    os.remove(tmp_path)
    return base64.b64encode(data)


def import_pdf(data):
    tmp_path = 'tmp.pdf'

    with open(tmp_path, 'wb') as f:
        f.write(data)

    f = open('tmp.pdf', 'rb')
    os.remove(tmp_path)

    try:
        reader = PyPDF2.PdfFileReader(f)
    except:
        return json.dumps({'body': 'ERR'})

    conn = db.connect('../db.sqlite')
    db.delete_from(conn, 'users')

    columns = db.get_columns(conn, 'users')[1:]
    data = {table: db.select(conn, table) for table in ['users', 'regions', 'cities']}

    for i in range(reader.getNumPages()):
        pdf = reader.getPage(i)

        lines = pdf.extractText().split('\n')
        full_name = lines[0].split(' ')

        values = {columns[i]: full_name[i] for i in range(len(full_name))}
        for j in range(1, len(lines)):
            values[columns[j + 2]] = lines[j].split(': ')[1]
        __apply([__name_to_id, __replace_invalid], values, data)
        __update(conn, columns, values, data)

    conn.close()
    return json.dumps({'body': 'OK'})


def export_pdf():
    tmp_path = 'tmp.pdf'

    conn = db.connect('../db.sqlite')
    data = {table: db.select(conn, table) for table in ['users', 'regions', 'cities']}
    conn.close()

    pdf = fpdf.FPDF(format='letter')
    pdf.add_font('DejaVu', '', '../res/DejaVuSansCondensed.ttf', uni=True)
    for i in range(len(data['users'])):
        __ids_to_names(data['users'][i], data)
        pdf.add_page()

        pdf.set_font('DejaVu', '', 20)
        pdf.cell(200, 20, txt='%s %s %s' % (data['users'][i]['second_name'],
                                            data['users'][i]['first_name'],
                                            data['users'][i]['patronymic']),
                 ln=1, align='C')

        pdf.set_font('DejaVu', '', 14)
        pdf.cell(200, 10, txt='Регион: %s' % data['users'][i]['region_id'], ln=1, align='L')
        pdf.cell(200, 10, txt='Город: %s' % data['users'][i]['city_id'], ln=2, align='L')
        pdf.cell(200, 10, txt='Контактный телефон: %s' % data['users'][i]['phone'], ln=3, align='L')
        pdf.cell(200, 10, txt='e-mail: %s' % data['users'][i]['email'], ln=4, align='L')

    pdf.output(tmp_path)
    with open(tmp_path, 'rb') as f:
        data = f.read()
        f.close()
    os.remove(tmp_path)
    return base64.b64encode(data)
